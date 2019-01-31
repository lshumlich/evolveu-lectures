
import sys
from openpyxl import load_workbook
# from openpyxl import Workbook
from datetime import datetime


def to_date(dt):
    return datetime.strptime(dt, "%Y-%m-%d").date()


def cell_has_value(ws, row, col):
    if row > ws.max_row:
        return False
    if col > ws.max_column:
        return False
    if not ws.cell(row=row, column=col).value:
        return False
    if not ws.cell(row=row, column=col).value.strip():
        return False

    return True


def find_tab(wb, date_s):
    for sheet in wb:
        if sheet.title == date_s[0:7]:
            return sheet
    return None


def find_date_row(ws, date_s):
    our_date = to_date(date_s)
    day = our_date.day
    # bad assumption that the 1st of the month is in row 2...
    # when we learn exceptions we will deal with this
    return day + 1


def find_available_spots(ws, row):
    spots = []
    if ws:
        for col in range(2, ws.max_column + 1):
            if not cell_has_value(ws, row, col):
                spots.append(ws.cell(row=1, column=col).value)
    return spots


# print('-->', c, ws.cell(row = row, column = c).value)

def find_available_spots_for_date(wb, date_s):
    ws = find_tab(wb, date_s)
    row = find_date_row(ws, date_s)
    return find_available_spots(ws, row)


def open_booking_wb(name):
    return load_workbook(name)


def process(args):
    # print('From process ---> The Args are:', args)
    print('Hello world from', __name__, __file__)

    if len(args) < 2:
        print('Please pass a date in the format of yyyy-mm-dd')
        return

    date = args[1]
    if len(args) > 2:
        file_name = args[2]
    else:
        file_name = 'cSpace_Booking.xlsx'

    print('Rooms Available for', date, 'in', file_name)

    for room in find_available_spots_for_date(open_booking_wb(file_name), date):
        print('-->', room)


if __name__ == '__main__':
    process(sys.argv)
