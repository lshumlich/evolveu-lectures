
import datetime
import unittest
from openpyxl import Workbook, worksheet
# from openpyxl import load_workbook

import tdd_excel


class TestTddExcel(unittest.TestCase):

    def test_to_date(self):
        dt = tdd_excel.to_date("2010-08-02")
        self.assertIsInstance(dt, datetime.date)
        self.assertEqual(2010, dt.year)
        self.assertEqual(8, dt.month)
        self.assertEqual(2, dt.day)

        dt = tdd_excel.to_date("2020-09-03")
        self.assertEqual(2020, dt.year)
        self.assertEqual(9, dt.month)
        self.assertEqual(3, dt.day)

    def test_cell_has_value(self):
        wb = Workbook()
        ws = wb.active

        self.assertFalse(tdd_excel.cell_has_value(ws, 1, 1))

        ws.cell(row=1, column=1).value = ''
        self.assertFalse(tdd_excel.cell_has_value(ws, 1, 1))

        ws.cell(row=1, column=1).value = 'Room a'
        self.assertTrue(tdd_excel.cell_has_value(ws, 1, 1))

        ws.cell(row=1, column=1).value = '   '
        self.assertFalse(tdd_excel.cell_has_value(ws, 1, 1))

        self.assertFalse(tdd_excel.cell_has_value(ws, 100, 100))
        self.assertFalse(tdd_excel.cell_has_value(ws, 1, 100))
        self.assertFalse(tdd_excel.cell_has_value(ws, 100, 1))

    def test_find_tab(self):
        wb = self.create_test_wb()
        entered_date = '2012-07-01'

        ws = tdd_excel.find_tab(wb, entered_date)
        self.assertIsInstance(ws, worksheet.Worksheet)
        self.assertEqual(entered_date[0:7], ws.title)

        self.assertIsNone(tdd_excel.find_tab(wb, '2018-07-01'))

    def test_find_date_row(self):
        wb = self.create_test_wb()
        entered_date = '2012-07-10'
        ws = tdd_excel.find_tab(wb, entered_date)

        self.assertEqual(2, tdd_excel.find_date_row(ws, '2012-07-01'))
        self.assertEqual(11, tdd_excel.find_date_row(ws, '2012-07-10'))
        self.assertEqual(29, tdd_excel.find_date_row(ws, '2012-07-28'))

    def test_find_available_spots(self):
        wb = self.create_test_wb()
        ws = wb["2012-07"]

        self.assertEqual(['Room c'],
                         tdd_excel.find_available_spots(ws, 2))

        self.assertEqual(['Room b', 'Room c'],
                         tdd_excel.find_available_spots(ws, 5))

        self.assertEqual([],
                         tdd_excel.find_available_spots(None, 5))

    def test_find_available_spots_for_date(self):
        wb = self.create_test_wb()

        self.assertEqual(['Room c'],
                         tdd_excel.find_available_spots_for_date(wb, "2012-07-01"))

        self.assertEqual(['Room b', 'Room c'],
                         tdd_excel.find_available_spots_for_date(wb, "2012-07-04"))

        self.assertEqual([],
                         tdd_excel.find_available_spots_for_date(wb, "2025-07-04"))

    def test_process(self):
        tdd_excel.process(['moduleName'])
        tdd_excel.process(['moduleName', '2018-07-01'])
        tdd_excel.process(['moduleName', '2012-07-04', 'test.xlsx'])

    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Mysheet")
        wb.create_sheet("Mysheet2")
        wb.create_sheet("2012-07")
        wb.create_sheet("2012-08")
        wb.create_sheet("2012-09")

        ws = wb["2012-07"]
        ws.cell(row=1, column=2).value = 'Room a'
        ws.cell(row=1, column=3).value = 'Room b'
        ws.cell(row=1, column=4).value = 'Room c'

        ws.cell(row=2, column=2).value = 'Harry'
        ws.cell(row=2, column=3).value = 'Harry'
        ws.cell(row=2, column=4).value = None

        ws.cell(row=5, column=2).value = 'Harry'

        # wb.save('test.xlsx')

        return wb
